# -*- coding:utf-8 -*-

from appium import webdriver
from appium.options.common.base import AppiumOptions
from appium.webdriver.common.touch_action import TouchAction
from appium.webdriver.common.multi_action import MultiAction
from appium import webdriver
from appium.webdriver.common.touch_action import TouchAction
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from openpyxl.utils import column_index_from_string
import subprocess


DosyaUzantisi = "<DosyaUzantisiAccount>"  # Placeholder for the file path of the social media account Excel file

def ChooseAccountForSharing(SocialMediaSelected="Instagram", Story=None):

    DosyaUzantisiSocial = "<DosyaUzantisiAccount>"  # Placeholder for the file path where social media account information is stored
    workbook = openpyxl.load_workbook(DosyaUzantisiSocial)
    worksheet = workbook["Sayfa1"]

    DosyaUzantisiForStory = "<DosyaUzantisiAccount>"  # Placeholder for the file path of the Excel file containing story data
    workbookForStory = openpyxl.load_workbook(DosyaUzantisiForStory)
    worksheetForStoryList = workbookForStory.sheetnames


    InstagramAccounts,SocialMedia,i = "Default","Default",0
    InstagramAccountsList=[]
    InstagramStoryList=worksheetForStoryList
    SutunList=["E","G","I","K","N","O"]
    SocialMediasList=[]


    while SocialMedia is not None:
        SocialMedia = worksheet[chr(ord("E")+2*i) + str(2)].value
        if SocialMedia is None:
            break
        i += 1
        SocialMediasList.append(SocialMedia)

    for SutunNo,OneSocialMedia in enumerate(SutunList):
        if SocialMediaSelected == worksheet[chr(ord(OneSocialMedia)) + str(2)].value:
            SutunNo=SutunNo
            break

    i=0
    while InstagramAccounts is not None:
        InstagramAccounts = worksheet[SutunList[SutunNo] + str(4 + i)].value
        if InstagramAccounts is None:
            break
        i += 1
        InstagramAccountsList.append(InstagramAccounts)

    print(f"Toplam Hesap Sayısı: {i}")

    for index,AccountsOne in enumerate(InstagramStoryList):
        print(f"{index+1} : {AccountsOne}")

    HesapNoSocial=input("Lütfen buraya hangi hesap ile işlem yapmak istediğinizi yazınız: ")
    i=0
    for index,AccountsOne in enumerate(InstagramStoryList,start=1):
        if index == int(HesapNoSocial):
            print(f"{index} : {AccountsOne}")
            UserNameSecilen = str((AccountsOne.split())[0])
            break

    while True:
        if str(UserNameSecilen)==str(worksheet[SutunList[SutunNo] + str(4 + i)].value):
            UserName=str(worksheet[SutunList[SutunNo] + str(4 + i)].value)
            Password=str(worksheet[chr(ord(SutunList[SutunNo])+1) + str(4 + i)].value)
            return UserName,Password,AccountsOne,UserNameSecilen
            break
        i +=1
UserName,Password,AccountsOne,UserNameSecilen=ChooseAccountForSharing()

NerdeKaldik=int(input("Nerede Kaldik,Hangi hikayede kaldık. En son tamamlanan hikaye :"))

if NerdeKaldik > 0:
    BaslangicDeger=int((NerdeKaldik+1)*4)
else:
    BaslangicDeger=int(4)

InstagramHesapAdi = AccountsOne
InstagramAdi=UserNameSecilen

CountForQuestions=5
workbook = openpyxl.load_workbook(DosyaUzantisi)
worksheet = workbook[InstagramHesapAdi]

def turkish_upper(text):
    turkish_letters = {'i': 'İ', 'ı': 'I', 'ğ': 'Ğ', 'ü': 'Ü', 'ş': 'Ş', 'ö': 'Ö', 'ç': 'Ç'}
    return ''.join(turkish_letters.get(c, c.upper()) for c in text)

def SoruCevapEtiketi(SoruNumarasi):
    CountForQuestions = 5
    ChooseABCD = []
    IkiSoruArasiMesafe = 8
    SorununKendisi= worksheet["B" + str(CountForQuestions + 1 + IkiSoruArasiMesafe*(int(SoruNumarasi)-1))].value
    for index, ChooseABCDone in enumerate(range(4)):
        SoruBaslangic = CountForQuestions + 3

        ABCD = worksheet["C" + str(SoruBaslangic + index + IkiSoruArasiMesafe*(int(SoruNumarasi)-1))].value
        ChooseABCD.append(ABCD)

    #### Doğru Cevabı Belirle ####
    for index in range(4):
        SoruBaslangic = CountForQuestions + 3
        IkiSoruArasiMesafe = 8

        DogruCevap = worksheet["A" + str(SoruBaslangic + index+ IkiSoruArasiMesafe*(int(SoruNumarasi)-1))].value
        if DogruCevap == "Doğru Cevap:":
            SorununDogruCevabi = str(worksheet["C" + str(SoruBaslangic + index+IkiSoruArasiMesafe*(int(SoruNumarasi)-1))].value)
            break

    return SorununDogruCevabi, ChooseABCD,SorununKendisi

def AnketEtiketi(AnketNumarasi):
    CountForAnket = 136
    ChooseABCD = []
    IkiSoruArasiMesafe = 8
    SorununKendisi = worksheet["B" + str(CountForAnket + 1 + IkiSoruArasiMesafe * (int(AnketNumarasi) - 1))].value
    for index, ChooseABCDone in enumerate(range(4)):
        SoruBaslangic = CountForAnket + 3

        ABCD = worksheet["C" + str(SoruBaslangic + index + IkiSoruArasiMesafe * (int(AnketNumarasi) - 1))].value
        ChooseABCD.append(ABCD)

    return ChooseABCD, SorununKendisi

def LinkEtkiketi(LinkNumarasi):
    CountForLink = 40
    İkiLinkArasindaki = 4
    LinkinAdi= worksheet["B" + str(CountForLink + 1 + İkiLinkArasindaki*(int(LinkNumarasi)-1))].value
    LinkinAdresi= worksheet["B" + str(CountForLink + 2 + İkiLinkArasindaki*(int(LinkNumarasi)-1))].value
    return LinkinAdi,LinkinAdresi

def VideoVeGorselEkle(VideoNumarasi):
    CountForImages = 5
    İkiGorselArasiMesafe = 4
    GorselAdresi= worksheet["H" + str(CountForImages + 1 + İkiGorselArasiMesafe*(int(VideoNumarasi)-1))].value
    return GorselAdresi

def AciklamaYazisiEkle(AciklamaYazisiNumarasi):
    CountForAciklama = 40
    IkiAciklamaArasiMesafe = 4
    AciklamaBasligi= worksheet["H" + str(CountForAciklama + 1 + IkiAciklamaArasiMesafe*(int(AciklamaYazisiNumarasi)-1))].value


    return AciklamaBasligi

def EtkiketAdi(EtkiketNo):
    CountForEtkiket = 72
    IkiEtiketArasiMesafe = 4
    EtkiketBasligi= worksheet["B" + str(CountForEtkiket + 1 + IkiEtiketArasiMesafe*(int(EtkiketNo)-1))].value


    return EtkiketBasligi

def Bahset(BahsetNo):
    CountForBahset = 72
    IkiEtiketArasiMesafe = 4
    BirisindenBahset= worksheet["H" + str(CountForBahset + 1 + IkiEtiketArasiMesafe*(int(BahsetNo)-1))].value


    return BirisindenBahset

def SoruSorEkitketi(SoruSorNO):
    CountForSoruSor = 104
    IkiEtiketArasiMesafe = 4
    SoruSorBasliklari= worksheet["B" + str(CountForSoruSor + 1 + IkiEtiketArasiMesafe*(int(SoruSorNO)-1))].value


    return SoruSorBasliklari

def SliderPartEtiketi(SliderPartNO):
    CountForSliderPart = 104
    IkiEtiketArasiMesafe = 4
    SliderPartBaslik= worksheet["H" + str(CountForSliderPart + 1 + IkiEtiketArasiMesafe*(int(SliderPartNO)-1))].value

    return SliderPartBaslik

def GifPartEtiketi(GifNO):
    CountForGIF = 136
    IkiEtiketArasiMesafe = 4
    GIFPartSecimi= str(worksheet["H" + str(CountForGIF + 1 + IkiEtiketArasiMesafe*(int(GifNO)-1))].value)
    GIFPartSecimiSplitted=GIFPartSecimi.split()
    GIFPartSecimiText=GIFPartSecimiSplitted[0]
    GIFPartSecimiNo=GIFPartSecimiSplitted[2]
    return GIFPartSecimiText,GIFPartSecimiNo

def CheckPasifOrAktif(Sutun,Satir):
    try:
        PasifOrAktif= worksheet[chr(ord(Sutun) + 1)  + str(Satir)].value
    except:
        SutunNUmarasi = column_index_from_string(Sutun)
        PasifOrAktif = str(worksheet.cell(row=Satir, column=SutunNUmarasi+1).value)
        time.sleep(0.5)
    if PasifOrAktif == "Pasif":
        return False
    elif PasifOrAktif == "Aktif":
        return True
    else:
        print("Geçerli Bir Değer Bulunamadı. Durduruluyor...")
        print(PasifOrAktif)
        return False

def OneDriveStoryPartToGallery(TotalStory):

    time.sleep(2)
    OneDriveGiris = WebDriverWait(driver, timeout=15).until(EC.presence_of_element_located((By.XPATH, '<OneDriveIcon>')))
    # <OneDriveIcon>: XPath for OneDrive icon
    OneDriveGiris.click()

    try:
        time.sleep(2)
        driver.find_element(By.XPATH, '<TabbarAllFiles>').click()
        # <TabbarAllFiles>: XPath for tabbar all files icon
    except:
        time.sleep(2)
        driver.find_element(By.XPATH, '<FilesButton>').click()
        # <FilesButton>: XPath for Files button with specific value
    time.sleep(2)
    SosyalMedyaBotFolder = driver.find_element(By.XPATH, '<SosyalMedyaBOTFolder>')
    # <SosyalMedyaBOTFolder>: XPath for SosyalMedyaBOT folder
    SosyalMedyaBotFolder.click()

    time.sleep(2)
    Instagram = WebDriverWait(driver, timeout=15).until(EC.presence_of_element_located((By.XPATH, '<InstagramFolder>')))
    # <InstagramFolder>: XPath for Instagram folder
    Instagram.click()

    time.sleep(2)
    StoryKlasoru = WebDriverWait(driver, timeout=15).until(EC.presence_of_element_located((By.XPATH, '<StoryFolder>')))
    # <StoryFolder>: XPath for Story folder
    StoryKlasoru.click()

    time.sleep(2)
    InstagramHesabi = WebDriverWait(driver, timeout=15).until(EC.presence_of_element_located((By.XPATH, '<InstagramAccount>')))
    # <InstagramAccount>: XPath for Instagram account folder
    InstagramHesabi.click()

    Artis = 4
    suffixlist = ["mp4", "jpeg", "png", "jpg"]
    a = 5

    ReelsGorselVideoList = []

    for i in range(BaslangicDeger, 36, 4):
        if CheckPasifOrAktif("N", i):
            VideoOrGorsel = str(worksheet[chr(ord("N")) + str(i)].value)
            HikayeSiralamasi = str(worksheet[chr(ord("N") - 1) + str(i)].value)
            HikayeSiralamasiSplitted = HikayeSiralamasi.split()
            HikayeSiralamasi = int(HikayeSiralamasiSplitted[2][1])
            ReelsGorselVideoList.append((HikayeSiralamasi, VideoOrGorsel))

        video_gorsel_dict = {}
        for eleman in ReelsGorselVideoList:
            if eleman[1] not in video_gorsel_dict:
                video_gorsel_dict[eleman[1]] = len(video_gorsel_dict) + 1

    numaralandirilmis_liste = [(video_gorsel_dict[eleman[1]], eleman[1]) for eleman in ReelsGorselVideoList]
    sadece_rakamlar = [int(str(eleman[0])[0]) for eleman in numaralandirilmis_liste]

    sirali_liste = []
    tekil = set()

    for eleman in numaralandirilmis_liste:
        if eleman[1] not in tekil:
            sirali_liste.append(eleman[1])
            tekil.add(str(eleman[1]))

    time.sleep(3)
    for ReelsGorselVideoListOne in sirali_liste:
        ReelsGorselVideoListOneSplitted = ReelsGorselVideoListOne.split()
        StoryListNumber = int(ReelsGorselVideoListOneSplitted[3][1])
        for a in range(5, 38, 4):
            try:
                FileNameUpper = str(worksheet[chr(ord("H")) + str(a)].value)
                FileNameUpperSplitted = FileNameUpper.split()
                FileNameNo = int(FileNameUpperSplitted[3][1])
                if StoryListNumber == FileNameNo:
                    UploadStory = str(worksheet[chr(ord("H")) + str(a + 1)].value)
                    time.sleep(2)
                    for x in suffixlist:
                        try:
                            # ikisi de çalışabiliyor...
                            try:
                                time.sleep(1)
                                driver.find_element(By.XPATH, '<FileUpload1>').click()
                                # <FileUpload1>: XPath for file upload option 1
                            except:
                                time.sleep(1)
                                driver.find_element(By.XPATH, '<FileUpload2>').click()
                                # <FileUpload2>: XPath for file upload option 2
                            time.sleep(0.5)
                        except:
                            pass
                    time.sleep(2)
                    driver.find_element(By.XPATH, '<DownloadButton>').click()
                    # <DownloadButton>: XPath for Download button
                    time.sleep(20)
                    driver.find_element(By.XPATH, '<BackButton>').click()
                    # <BackButton>: XPath for Back button
                    time.sleep(2)
                    break
            except:
                pass

    for index, Object in enumerate(sirali_liste):
        ForImagesAdjust = index + 1
    return sadece_rakamlar, ForImagesAdjust


options = AppiumOptions()
options.load_capabilities({
    "platformName": "<PlatformName>",
    # <PlatformName>: Name of the platform (e.g., "iOS")
    "appium:udid": "<UDID>",
    # <UDID>: Unique Device Identifier
    "appium:automationName": "<AutomationName>",
    # <AutomationName>: Name of the automation engine (e.g., "XCUITest")
    "appium:deviceName": "<DeviceName>",
    # <DeviceName>: Name of the device (e.g., "iPad")
    "appium:includeSafariInWebviews": "<IncludeSafariInWebviews>",
    # <IncludeSafariInWebviews>: Boolean to include Safari in webviews
    "appium:newCommandTimeout": "<NewCommandTimeout>",
    # <NewCommandTimeout>: Timeout for new commands (e.g., 3600)
    "appium:connectHardwareKeyboard": "<ConnectHardwareKeyboard>"
    # <ConnectHardwareKeyboard>: Boolean to connect hardware keyboard
})

driver = webdriver.Remote("<RemoteURL>", options=options)
# <RemoteURL>: URL of the remote Appium server (e.g., "http://125.0.1:2583/wd/hub")

def ZoomInOrOut(element, ikiParmakArasiMesafe=30, ParmaklariYaklastir=10, YakinlastirOrUzaklasitir="Yakinlastir", Sutun="N", Satir=4):
    try:
        KacXZoomOlsunExcel = worksheet[chr(ord(Sutun)) + str(Satir+2)].value
    except:
        Sutun = column_index_from_string(Sutun)
        KacXZoomOlsunExcel = worksheet.cell(row=Satir+2, column=Sutun).value

    KacXZoomOlsunSplitted = KacXZoomOlsunExcel.split()
    KacXZoomOlsun = int(KacXZoomOlsunSplitted[0][0])

    for i in range(KacXZoomOlsun):
        if YakinlastirOrUzaklasitir == "Yakinlastir":
            ParmaklariYaklastir = ParmaklariYaklastir
        elif YakinlastirOrUzaklasitir == "Uzaklastir":
            ParmaklariYaklastir = -ParmaklariYaklastir
        else:
            print("Lütfen Geçerli Bir Değer Giriniz:")
        # Elementin konumunu al

        # 374,272
        # Ekran boyutlarını al
        window_size = driver.get_window_size()
        screen_width = window_size['width']
        screen_height = window_size['height']
        # 375,667

        # Elementin boyutunu al
        size = element.size
        location = element.location

        x_coordinate = location['x'] + size['width'] / 2
        y_coordinate = location['y'] + size['height'] / 2

        touch_action2 = TouchAction(driver)
        touch_action2.press(None, x=x_coordinate+ikiParmakArasiMesafe, y=y_coordinate).wait(1000).move_to(x=x_coordinate+ikiParmakArasiMesafe-ParmaklariYaklastir, y=y_coordinate).wait(1000).release()
        time.sleep(1)
        # Üçüncü TouchAction
        touch_action3 = TouchAction(driver)
        touch_action3.press(None, x=x_coordinate-ikiParmakArasiMesafe, y=y_coordinate).wait(1000).move_to(x=x_coordinate-ikiParmakArasiMesafe+ParmaklariYaklastir, y=y_coordinate).wait(1000).release()  # 13

        # MultiAction oluştur
        multi_action = MultiAction(driver)
        multi_action.add(touch_action3, touch_action2)
        time.sleep(1)
        multi_action.perform()

def zoomTextSpecial(element, Sutun, Satir):
    element.click()
    time.sleep(1)
    KacXZoomOlsunExcel = worksheet[chr(ord(Sutun)) + str(Satir+2)].value
    KacXZoomOlsunExcelSplitted = KacXZoomOlsunExcel.split()
    KacXZoomOlsunExcel = KacXZoomOlsunExcelSplitted[3][0]
    for i in range(int(KacXZoomOlsunExcel)):
        # Ekran boyutlarını al
        window_size = driver.get_window_size()
        screen_width = window_size['width']
        screen_height = window_size['height']
        # Elementin boyutunu al
        touch_action2 = TouchAction(driver)
        touch_action2.press(x=120, y=400).wait(1000).move_to(x=120, y=450).release().perform()
        time.sleep(0.5)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<TextEntryDoneButton>'))).click()
    # <TextEntryDoneButton>: XPath for text entry done button

def linkekle(Link, cikartmametni):
    time.sleep(2)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AddStickersButton>'))).click()
    # <AddStickersButton>: XPath for add stickers button
    time.sleep(3)
    driver.swipe(430, 500, 430, 200, duration=200)
    time.sleep(3)
    StickerName = '<LinkStickerTrayCell>'
    # <LinkStickerTrayCell>: XPath for link sticker tray cell
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, StickerName))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<TextFieldPlaceholder>'))).send_keys(Link)
    # <TextFieldPlaceholder>: XPath for text field placeholder
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<CustomizeStickerText>'))).click()
    # <CustomizeStickerText>: XPath for customize sticker text
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<StickerText>'))).send_keys(cikartmametni)
    # <StickerText>: XPath for sticker text
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<DoneButton>'))).click()
    # <DoneButton>: XPath for done button

def KonuEtkiketiEkle(Hashtag):
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AddStickersButton>'))).click()
    # <AddStickersButton>: XPath for add stickers button
    StickerName = '<HashtagStickerTrayCell>'
    # <HashtagStickerTrayCell>: XPath for hashtag sticker tray cell
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, StickerName))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<HashtagTextView>'))).send_keys(Hashtag)
    # <HashtagTextView>: XPath for hashtag text view
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<DoneButton>'))).click()
    # <DoneButton>: XPath for done button

def BahsetEtiketle(Birisi):
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AddStickersButton>'))).click()
    # <AddStickersButton>: XPath for add stickers button
    StickerName = '<MentionStickerTrayCell>'
    # <MentionStickerTrayCell>: XPath for mention sticker tray cell
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, StickerName))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<MentionSticker>'))).send_keys(Birisi)
    # <MentionSticker>: XPath for mention sticker
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<DoneButton>'))).click()
    # <DoneButton>: XPath for done button

def SorularEtkiketiEkle(SoruBasligi):
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AddStickersButton>'))).click()
    # <AddStickersButton>: XPath for add stickers button
    StickerName = '<QuestionSticker>'
    # <QuestionSticker>: XPath for question sticker
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, StickerName))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<QuestionTextField>'))).send_keys(SoruBasligi)
    # <QuestionTextField>: XPath for question text field
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<DoneButton>'))).click()
    # <DoneButton>: XPath for done button

def SliderEtiketiEkle(SliderBasligi):
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AddStickersButton>'))).click()
    # <AddStickersButton>: XPath for add stickers button
    StickerName = '<SliderStickerTrayCell>'
    # <SliderStickerTrayCell>: XPath for slider sticker tray cell
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, StickerName))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<SliderTextView>'))).send_keys(SliderBasligi)
    # <SliderTextView>: XPath for slider text view
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<DoneButton>'))).click()
    # <DoneButton>: XPath for done button




def GIFEtiketEkle(GIFPartSecimiText, GIFPartSecimiNo):
    # This function is used to insert a specific GIF tag.
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AddStickersButton>'))).click()
    # <AddStickersButton>: XPath for add stickers button
    StickerName = '<GIFStickerTrayCell>'
    # <GIFStickerTrayCell>: XPath for GIF sticker tray cell
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, StickerName))).click()

    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<SearchTextInput>'))).clear()
        # <SearchTextInput>: XPath for search text input
    except:
        pass
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<SearchTextInput>'))).send_keys(GIFPartSecimiText)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<GIFSelection>[' + str(GIFPartSecimiNo) + ']'))).click()
    # <GIFSelection>: XPath for GIF selection

def CloseGIFTab():
    # This function is used to close the GIF tag tab.
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AddStickersButton>'))).click()
    # <AddStickersButton>: XPath for add stickers button
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<CancelButton>'))).click()
    # <CancelButton>: XPath for cancel button
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AdjustSizeButton>'))).click()
    # <AdjustSizeButton>: XPath for adjust size button

def Konumlandir(element, Sutun, Satir):
    # This function is used to move an element to a specific position.
    try:
        YTasi = worksheet[chr(ord(Sutun) + 1) + str(Satir+1)].value
        XTasi = worksheet[chr(ord(Sutun)) + str(Satir+1)].value
    except:
        Sutun = column_index_from_string(Sutun)
        YTasi = worksheet.cell(row=Satir+1, column=Sutun+1).value
        XTasi = worksheet.cell(row=Satir+1, column=Sutun).value
    time.sleep(1)

    YTasiSplitted = YTasi.split(":")
    YTasi = int(YTasiSplitted[1])

    XTasiSplitted = XTasi.split(":")
    XTasi = int(XTasiSplitted[1])

    action = TouchAction(driver)
    time.sleep(1)

    size = element.size
    location = element.location

    x_coordinate = location['x'] + size['width'] / 2
    y_coordinate = location['y'] + size['height'] / 2

    time.sleep(1)
    action.press(x=x_coordinate, y=y_coordinate).wait(1000).move_to(x=XTasi, y=YTasi).release().perform()

def renksec(element, Sutun, Satir, A, RenkDegeri="White", LinkinAdi="0"):
    # This function is used to replace a certain element with a certain colour or a certain text.
    if str(A) == "Soru" or str(A) == "SoruSor" or str(A) == "Anket":
        element.click()
        try:
            x = worksheet[chr(ord(Sutun) + 1) + str(Satir + 2)].value
        except:
            Sutun = column_index_from_string(Sutun)
            x = worksheet.cell(row=Satir + 2, column=Sutun + 1).value

        xSplitted = x.split(" - ")
        x = int(xSplitted[0])

        for i in range(x):
            driver.find_element(By.XPATH, '<ToggleBackgroundColorButton>').click()
            # <ToggleBackgroundColorButton>: XPath for toggle background color button
            time.sleep(0.5)
        try:
            driver.find_element(By.XPATH, '<DoneButton>').click()
            # <DoneButton>: XPath for done button
        except:
            driver.find_element(By.XPATH, '<DoneButtonAlt>').click()
            # <DoneButtonAlt>: Alternative XPath for done button
    elif str(A) == "Açıklama":
        element.click()
        RenkDegeri = str(worksheet[chr(ord(Sutun) + 1) + str(Satir + 2)].value)
        RenkDegeriSplitted = RenkDegeri.split()
        RenkDegeri = str(RenkDegeriSplitted[2])

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<TextColorButton>'))).click()
        # <TextColorButton>: XPath for text color button
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<ColorSelection>[@value="' + str(RenkDegeri) + '"]'))).click()
        # <ColorSelection>: XPath for color selection
        driver.find_element(By.XPATH, '<DoneButtonAlt>').click()
        # <DoneButtonAlt>: Alternative XPath for done button

    elif str(A) == "Link":
        try:
            x = worksheet[chr(ord(Sutun) + 1) + str(Satir + 2)].value
        except:
            Sutun = column_index_from_string(Sutun)
            x = worksheet.cell(row=Satir + 2, column=Sutun + 1).value
        xSplitted = x.split(" - ")
        x = int(xSplitted[0])

        for i in range(x):
            try:
                StickerName = '//XCUIElementTypeStaticText[@name="' + str(turkish_upper(LinkinAdi)) + '"]'
                StickerNameXpath = driver.find_element(By.XPATH, StickerName)
            except:
                StickerName = '//XCUIElementTypeStaticText[@name="' + str(LinkinAdi).capitalize() + '"]'
                StickerNameXpath = driver.find_element(By.XPATH, StickerName)

            StickerNameXpath.click()
            time.sleep(0.5)

    elif str(A) == "Etiket" or str(A) == "Bahset":
        try:
            x = worksheet[chr(ord(Sutun) + 1) + str(Satir + 2)].value
        except:
            Sutun = column_index_from_string(Sutun)
            x = worksheet.cell(row=Satir + 2, column=Sutun + 1).value
        xSplitted = x.split(" - ")
        x = int(xSplitted[0])

        for i in range(x):
            element.click()
            time.sleep(0.5)
    else:
        print("Error")

def SoruCevapEkle(BaslikAdi="test", A="Default", B="Default", C="Default1", D="Default", DogruCevapHangisi="Default1"):
    # This function is used to add a question-answer tag.
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AddStickersButton>'))).click()
    # <AddStickersButton>: XPath for add stickers button
    StickerName = '<QuizStickerTrayCell>'
    # <QuizStickerTrayCell>: XPath for quiz sticker tray cell

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, StickerName))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<QuizSticker>'))).click()
    # <QuizSticker>: XPath for quiz sticker

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<QuizTextEntry>'))).send_keys(BaslikAdi)
    # <QuizTextEntry>: XPath for quiz text entry

    SenecekSayisilari = [A, B]
    for index, SenecekSayisi in enumerate(SenecekSayisilari):
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<ChoiceTextEntry>[' + str(index + 1) + ']'))).send_keys(SenecekSayisi)
        # <ChoiceTextEntry>: XPath for choice text entry

    SenecekSayisilari = [C, D]
    for index, SecenekSayisi in enumerate(SenecekSayisilari):
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AddChoiceTextEntry>'))).send_keys(SecenekSayisi)
        # <AddChoiceTextEntry>: XPath for add choice text entry

    time.sleep(1)
    DogruCevap = str(DogruCevapHangisi)
    DogruCevapXPath = '<CorrectAnswerSelection>[' + DogruCevap + ']'
    # <CorrectAnswerSelection>: XPath for correct answer selection

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, DogruCevapXPath))).click()
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<DoneButton>'))).click()
        # <DoneButton>: XPath for done button
    except:
        driver.find_element(By.XPATH, '<DoneButtonAlt>').click()
        # <DoneButtonAlt>: Alternative XPath for done button

def AnketEtiketiEkle(BaslikAdi, A, B, C, D):
    # This function is used to add a survey tag.
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AddStickersButton>'))).click()
    # <AddStickersButton>: XPath for add stickers button

    try:
        StickerName = '<PollStickerTrayCell>'
        # <PollStickerTrayCell>: XPath for poll sticker tray cell
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, StickerName))).click()
    except:
        StickerName = '<AlternativePollStickerTrayCell>'
        # <AlternativePollStickerTrayCell>: Alternative XPath for poll sticker tray cell
        WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, StickerName))).click()

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<PollTextEntry>'))).send_keys(BaslikAdi)
    # <PollTextEntry>: XPath for poll text entry
    time.sleep(5)
    SenecekSayisilari = [A, B]
    YesOrNO = ["Evet", "Hayır"]
    for index, SenecekSayisi in enumerate(SenecekSayisilari):
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<PollOption>[' + YesOrNO[int(index)] + ']'))).send_keys(SenecekSayisi)
        # <PollOption>: XPath for poll option

    SenecekSayisilari = [C, D]
    for index, SecenekSayisi in enumerate(SenecekSayisilari):
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AddPollOption>'))).send_keys(SecenekSayisi)
        # <AddPollOption>: XPath for add poll option

    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<DoneButton>'))).click()
        # <DoneButton>: XPath for done button
    except:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<DoneButtonAlt>'))).click()
        # <DoneButtonAlt>: Alternative XPath for done button

def MetinEkle(Text, Sutun, Satir, MetinAnimasyonu="Pasif"):
    # This function is used to insert a specific text and set its style.
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AddTextButton>'))).click()
    # <AddTextButton>: XPath for add text button
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<TextEntryView>'))).send_keys(Text)
    # <TextEntryView>: XPath for text entry view
    YaziStiliNoArtis = 1
    YaziStiliNo = str(worksheet[(Sutun) + str(Satir + 3)].value)
    YaziStiliNoSplitted = YaziStiliNo.split()
    YaziStiliNo = int(YaziStiliNoSplitted[2])

    while int(YaziStiliNoArtis) <= int(YaziStiliNo):
        YaziStili = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '<TextStyle>[' + str(YaziStiliNoArtis) + ']')))
        # <TextStyle>: XPath for text style
        YaziStili.click()
        YaziStiliNoArtis += 1
        time.sleep(0.5)

    MetinVurgusuDurum = str(worksheet[chr(ord(Sutun) + 1) + str(Satir + 3)].value)
    MetinVurgusuDurumSplitted = MetinVurgusuDurum.split()
    MetinVurgusuDurum = str(MetinVurgusuDurumSplitted[2])

    # Devre Dışı, Etkin, Ters
    MetinVurgusuList = ["Etkin", "Devre Dışı", "Ters"]
    while True:
        element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '<HighlightTextButton>')))
        # <HighlightTextButton>: XPath for highlight text button
        element.click()
        MetinVurgusuDurumxPath = str(element.get_attribute("value"))
        if MetinVurgusuDurumxPath == MetinVurgusuDurum:
            break

    if MetinAnimasyonu == "Aktif":
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<TextAnimationButton>'))).click()
        # <TextAnimationButton>: XPath for text animation button
    else:
        pass

    time.sleep(2)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<TextEntryDoneButton>'))).click()
    # <TextEntryDoneButton>: XPath for text entry done button

def KacTaneVar(EtkiketAdi, HarfGroup):
    # This function is used to calculate the number of times a particular tag has been used.
    ToplamEtiketSayisi = 0

    for i in range(4, 44, 4):
        for OneHarf in HarfGroup:
            try:
                CheckEtiket = str(worksheet[chr(ord(OneHarf)) + str(i)].value)
                CheckEtiketSplitted = CheckEtiket.split()
                EtiketCH = str(CheckEtiketSplitted[0])
                if EtiketCH == EtkiketAdi:
                    if CheckPasifOrAktif(OneHarf, i) == True:
                        ToplamEtiketSayisi += 1
            except:
                pass

    print(f"Kullanılan Toplam Hikaye Sayisi = {ToplamEtiketSayisi}")
    return int(ToplamEtiketSayisi)

def SelectImages(InstagramAdi, No, ToplamStory, Sutun, Satir):
    # This function is used to select images from a specific Instagram story.
    time.sleep(3)
    StoryAdd = str(worksheet[chr(ord(Sutun) - 1) + str(Satir)].value)
    StoryAddSplitted = StoryAdd.split()
    StoryAddNo = int(StoryAddSplitted[2][1])

    VideoGorsel = str(worksheet[chr(ord(Sutun)) + str(Satir)].value)
    VideoNumarasiNoSplitted = VideoGorsel.split()
    VideoNumarasiNo = int(VideoNumarasiNoSplitted[3][1])
    time.sleep(3)
    Parameters = 0
    try:
        Parameters = 3
        time.sleep(2)
        WebDriverWait(driver, timeout=5).until(EC.presence_of_element_located((By.XPATH, '<StoryTrayCell>[' + str(Parameters) + ']'))).click()
        # <StoryTrayCell>: XPath for story tray cell
        try:
            time.sleep(2)
            WebDriverWait(driver, timeout=5).until(EC.presence_of_element_located((By.XPATH, '<StoryCameraGalleryButton>'))).click()
            # <StoryCameraGalleryButton>: XPath for story camera gallery button
        except:
            pass
    except:
        time.sleep(2)
        Parameters = 2
        WebDriverWait(driver, timeout=5).until(EC.presence_of_element_located((By.XPATH, '<StoryTrayCell>[' + str(Parameters) + ']'))).click()
        time.sleep(2)
        try:
            WebDriverWait(driver, timeout=5).until(EC.presence_of_element_located((By.XPATH, '<StoryCameraGalleryButton>'))).click()
            # <StoryCameraGalleryButton>: XPath for story camera gallery button
        except:
            pass

    time.sleep(3)
    driver.find_element(By.XPATH, '<SelectButton>').click()
    # <SelectButton>: XPath for select button
    # No = str(ToplamStory - StoryAddNo)
    time.sleep(3)
    ChooseImagesForSharing = '<GalleryVideoCellPrefix>' + str(No) + '<GalleryVideoCellSuffix>'
    # <GalleryVideoCellPrefix>: XPath prefix for gallery video cell
    # <GalleryVideoCellSuffix>: XPath suffix for gallery video cell
    try:
        time.sleep(1)
        driver.find_element(By.XPATH, ChooseImagesForSharing).click()
        time.sleep(2)
    except:
        try:
            driver.swipe(430, 700, 430, 500, duration=200)
            time.sleep(1)
            driver.find_element(By.XPATH, ChooseImagesForSharing).click()
        except:
            driver.swipe(430, 600, 430, 1000, duration=200)
            time.sleep(1)
            driver.find_element(By.XPATH, ChooseImagesForSharing).click()

    time.sleep(2)
    WebDriverWait(driver, timeout=5).until(EC.presence_of_element_located((By.XPATH, '<GallerySelectionNextButton>'))).click()
    # <GallerySelectionNextButton>: XPath for gallery selection next button
    try:
        time.sleep(1)
        driver.find_element(By.XPATH, '<DoneButtonAlt>').click()
        # <DoneButtonAlt>: Alternative XPath for done button
        time.sleep(2)
    except:
        pass


HarfGroup = ["N", "P", "R", "T", "V", "X", "Z", "AB", "AD", "AF"]

driver.background_app(-1)
time.sleep(2)

ToplamStory = KacTaneVar("Video", HarfGroup)

sadece_rakamlar, ForImagesAdjust = OneDriveStoryPartToGallery(ToplamStory)

time.sleep(2)
driver.background_app(-1)
WebDriverWait(driver, timeout=5).until(EC.presence_of_element_located((By.XPATH, '<InstagramIcon>'))).click()
# <InstagramIcon>: XPath for Instagram icon

# Find the element
element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<ProfileTabButton>')))
# <ProfileTabButton>: XPath for profile tab button

# Long press on the element
action = TouchAction(driver)
time.sleep(2)
action.long_press(element).wait(1000).release().perform()

try:
    xpath_expression = "(//XCUIElementTypeButton[contains(@name, '"+str(InstagramAdi.lower()) +"')])[1]"
    WebDriverWait(driver, timeout=5).until(EC.presence_of_element_located((By.XPATH, xpath_expression))).click()
except:
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<AddInstagramAccountButton>'))).click()
    # <AddInstagramAccountButton>: XPath for add Instagram account button
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<LoginButton>'))).click()
    # <LoginButton>: XPath for login button
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<UsernameField>'))).send_keys(UserName)
    # <UsernameField>: XPath for username field
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<PasswordField>'))).send_keys(Password)
    # <PasswordField>: XPath for password field
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<LoginButton>'))).click()
    # <LoginButton>: XPath for login button

try:
    time.sleep(1)
    driver.find_element(By.XPATH, '<FeedControlsMenuButton>').click()
    # <FeedControlsMenuButton>: XPath for feed controls menu button
except:
    pass

WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '<MainFeedTabButton>'))).click()
# <MainFeedTabButton>: XPath for main feed tab button

try:
    for index, i in enumerate(range(BaslangicDeger, 36, 4), start=0): 
        try:
            for Harf in HarfGroup:
                Etiket = worksheet[Harf+str(i)].value
                try:
                    EtiketParts = Etiket.split()
                    A = EtiketParts[0]
                except:
                    A = "Deleted"

                if A == "Video":
                    if CheckPasifOrAktif(Harf, i) == True:
                        VideoNumarasi = EtiketParts[3][1]
                        GorselAdresi = VideoVeGorselEkle(VideoNumarasi)
                        time.sleep(3)
                        No = str(int(ForImagesAdjust) - int(sadece_rakamlar[index]))
                        SelectImages(InstagramHesapAdi, No=No, ToplamStory=ForImagesAdjust, Sutun=Harf, Satir=i)
                    if CheckPasifOrAktif(Harf, i) == False:
                        break

                elif str(A) == "Soru":
                    if CheckPasifOrAktif(Harf, i) == True:
                        SoruNumarasi = EtiketParts[1][1:]
                        dogru_cevap, choose_ABCD, SorununKendisi = SoruCevapEtiketi(SoruNumarasi)
                        SoruCevapEkle(BaslikAdi=SorununKendisi, A=choose_ABCD[0], B=choose_ABCD[1], C=choose_ABCD[2], D=choose_ABCD[3], DogruCevapHangisi=dogru_cevap)

                        SoruCevapDriver = driver.find_element(By.XPATH, '<QuizSticker>')
                        # <QuizSticker>: XPath for quiz sticker
                        ZoomInOrOut(SoruCevapDriver, Sutun=Harf, Satir=i)
                        Konumlandir(SoruCevapDriver, Sutun=Harf, Satir=i)
                        renksec(SoruCevapDriver, Sutun=Harf, Satir=i, A=A)
                    else:
                        pass

                elif A == "Link":
                    if CheckPasifOrAktif(Harf, i) == True:
                        LinkNumarasi = EtiketParts[1][1]
                        LinkinAdi, LinkinAdresi = LinkEtkiketi(LinkNumarasi)
                        linkekle(LinkinAdresi, LinkinAdi)

                        try:
                            StickerName = '<LinkSticker>[@name="'+str(turkish_upper(LinkinAdi))+'"]'
                            driver.find_element(By.XPATH, StickerName)
                        except:
                            StickerName = '<LinkSticker>[@name="'+str(LinkinAdi).capitalize()+'"]'
                            driver.find_element(By.XPATH, StickerName)

                        LinkDriver = driver.find_element(By.XPATH, StickerName)
                        ZoomInOrOut(LinkDriver, Sutun=Harf, Satir=i)
                        Konumlandir(LinkDriver, Sutun=Harf, Satir=i)
                        renksec(LinkDriver, Sutun=Harf, Satir=i, A=A, LinkinAdi=LinkinAdi)
                    else:
                        pass

                elif A == "Açıklama":
                    if CheckPasifOrAktif(Harf, i) == True:
                        AciklamaNumarasi = EtiketParts[2][1]
                        AciklamaBasligi = AciklamaYazisiEkle(AciklamaNumarasi)
                        MetinEkle(Text=AciklamaBasligi, Sutun=Harf, Satir=i)

                        TextDriver = driver.find_element(By.XPATH, '<DescriptionText>[@name="'+str(AciklamaBasligi)+'"]')
                        # <DescriptionText>: XPath for description text
                        zoomTextSpecial(TextDriver, Sutun=Harf, Satir=i)
                        ZoomInOrOut(TextDriver, Sutun=Harf, Satir=i)
                        Konumlandir(TextDriver, Sutun=Harf, Satir=i)
                        renksec(TextDriver, Sutun=Harf, Satir=i, A=A)
                    else:
                        pass

                elif A == "Etiket":
                    if CheckPasifOrAktif(Harf, i) == True:
                        KonuEtkinetNumarasi = EtiketParts[1][1]
                        KonuHashtagName = EtkiketAdi(KonuEtkinetNumarasi)
                        KonuEtkiketiEkle(KonuHashtagName)

                        StickerName = '<HashtagSticker>'
                        # <HashtagSticker>: XPath for hashtag sticker
                        EtiketDriver = driver.find_element(By.XPATH, StickerName)
                        ZoomInOrOut(EtiketDriver, Sutun=Harf, Satir=i)
                        Konumlandir(EtiketDriver, Sutun=Harf, Satir=i)
                        renksec(EtiketDriver, Sutun=Harf, Satir=i, A=A)

                elif A == "Bahset":
                    if CheckPasifOrAktif(Harf, i) == True:
                        BirisindenBahset = EtiketParts[1][1]
                        BahsedilenIsim = Bahset(BirisindenBahset)
                        BahsetEtiketle(BahsedilenIsim)

                        StickerName = '<MentionSticker>'
                        # <MentionSticker>: XPath for mention sticker
                        BahsetDriver = driver.find_element(By.XPATH, StickerName)
                        ZoomInOrOut(BahsetDriver, Sutun=Harf, Satir=i)
                        Konumlandir(BahsetDriver, Sutun=Harf, Satir=i)
                        renksec(BahsetDriver, Sutun=Harf, Satir=i, A=A)

                elif A == "SoruSor":
                    if CheckPasifOrAktif(Harf, i) == True:
                        SorSorPart = EtiketParts[1][1]
                        SorulanSoru = SoruSorEkitketi(SorSorPart)
                        SorularEtkiketiEkle(SorulanSoru)

                        StickerName = '<QuestionSticker>'
                        # <QuestionSticker>: XPath for question sticker
                        SoruSorDriver = driver.find_element(By.XPATH, StickerName)
                        ZoomInOrOut(SoruSorDriver, Sutun=Harf, Satir=i)
                        Konumlandir(SoruSorDriver, Sutun=Harf, Satir=i)
                        renksec(SoruSorDriver, Sutun=Harf, Satir=i, A=A)

                elif A == "Slider":
                    if CheckPasifOrAktif(Harf, i) == True:
                        SliderPart = EtiketParts[1][1]
                        SliderPartAdi = SliderPartEtiketi(SliderPart)
                        SliderPartEkle = SliderEtiketiEkle(SliderPartAdi)

                        StickerName = '<SliderSticker>'
                        # <SliderSticker>: XPath for slider sticker
                        SliderDriver = driver.find_element(By.XPATH, StickerName)
                        ZoomInOrOut(SliderDriver, Sutun=Harf, Satir=i)
                        Konumlandir(SliderDriver, Sutun=Harf, Satir=i)
                        renksec(SliderDriver, Sutun=Harf, Satir=i, A=A)

                elif A == "Anket":
                    if CheckPasifOrAktif(Harf, i) == True:
                        AnketNumarasi = EtiketParts[1][1]
                        ChooseABCD, SorununKendisi = AnketEtiketi(AnketNumarasi)
                        AnketEtiketiEkle(BaslikAdi=SorununKendisi, A=ChooseABCD[0], B=ChooseABCD[1], C=ChooseABCD[2], D=ChooseABCD[3])

                        AnketDriver = driver.find_element(By.XPATH, '<PollSticker>')
                        # <PollSticker>: XPath for poll sticker
                        ZoomInOrOut(AnketDriver, Sutun=Harf, Satir=i)
                        Konumlandir(AnketDriver, Sutun=Harf, Satir=i)
                        renksec(AnketDriver, Sutun=Harf, Satir=i, A=A)

                elif A == "Gif":
                    if CheckPasifOrAktif(Harf, i) == True:
                        GifNo = EtiketParts[1][1]
                        GIFPartSecimiText, GIFPartSecimiNo = GifPartEtiketi(GifNo)
                        GIFEtiketEkle(GIFPartSecimiText, GIFPartSecimiNo)

                        TotalGifNo = 1
                        GIFDriver = driver.find_element(By.XPATH, '<GIFSticker>[' + str(TotalGifNo) + ']')
                        # <GIFSticker>: XPath for GIF sticker
                        ZoomInOrOut(GIFDriver, Sutun=Harf, Satir=i)
                        Konumlandir(GIFDriver, Sutun=Harf, Satir=i)
                        CloseGIFTab()
                    else:
                        pass
                else:
                    pass
        except:
            pass

        try:
            driver.find_element(By.XPATH, '<ShareToStoryButton>').click()
            # <ShareToStoryButton>: XPath for share to story button
            KacinciPaylasimdayiz = worksheet["M" + str(i)].value
            print(f"{KacinciPaylasimdayiz} completed")
        except:
            driver.find_element(By.XPATH, '<YourStoriesButton>').click()
            # <YourStoriesButton>: XPath for your stories button
            KacinciPaylasimdayiz = worksheet["M" + str(i)].value
            print(f"{KacinciPaylasimdayiz} completed")
            time.sleep(1)
except:
    pass

print("Completed..")

driver.quit()
